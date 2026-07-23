"""Gas & shared-cost tracker.

Run this file with Python on Windows.  It stores its data locally in
gas_tracker.db and imports 407.xlsx once on first launch when that workbook is
in the same folder.  The original workbook is never edited.
"""

from __future__ import annotations

import calendar
import csv
from decimal import Decimal, ROUND_HALF_UP
from copy import copy
import json
import os
import shutil
import sqlite3
import subprocess
import sys
import threading
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from openpyxl import Workbook


APP_NAME = "Gas & Cost Splitter"
BASE_DIR = Path(getattr(sys, "_MEIPASS", Path(__file__).resolve().parent))
DATA_DIR = Path(__file__).resolve().parent
DB_PATH = DATA_DIR / "gas_tracker.db"
SOURCE_WORKBOOK = DATA_DIR / "407.xlsx"
MONEY = "$"
DEFAULT_PEOPLE = ["Abdul", "Uday", "Gurpreet"]
DEFAULT_DRIVERS = {"Uday": 8.2, "Gurpreet": 7.9}
PUBLISH_COMMAND = [sys.executable, str(DATA_DIR / "publish_readonly.py")]
GIT_PUBLISH_COMMAND = ["git", "add", "docs/index.html", "docs/.nojekyll"]
GIT_COMMIT_COMMAND = ["git", "commit", "-m", "Update read-only GitHub Pages snapshot"]
GIT_PUSH_COMMAND = ["git", "push", "origin", "master"]


def money(value: float) -> str:
    return f"{MONEY}{value:,.2f}"


def round_money(value: float) -> float:
    """Round monetary values once, using the normal financial half-up rule."""
    return float(Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def split_cents(total: float, attendees: list[str], payer: str) -> dict[str, float]:
    """Split an amount exactly to cents without short-changing the person who paid."""
    if not attendees:
        raise ValueError("At least one person must share an event.")
    cents = int(Decimal(str(round_money(total))) * 100)
    base, extra = divmod(cents, len(attendees))
    # When a cent cannot be shared evenly, charge the non-payers first. The
    # payer fronted the whole cost and should receive the larger reimbursement,
    # not be the person who absorbs that rounding penny.
    order = [person for person in attendees if person != payer] + ([payer] if payer in attendees else [])
    shares = {person: base for person in attendees}
    for person in order[:extra]:
        shares[person] += 1
    return {person: value / 100 for person, value in shares.items()}


def parse_date(value: str) -> str:
    """Return ISO date, accepting YYYY-MM-DD and common North American forms."""
    value = value.strip()
    for pattern in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
        try:
            return datetime.strptime(value, pattern).date().isoformat()
        except ValueError:
            pass
    raise ValueError("Use a date like 2026-07-14.")


class CalendarPopup(tk.Toplevel):
    """A compact, dependency-free date picker for the desktop forms."""
    def __init__(self, input_widget: "DateInput"):
        super().__init__(input_widget)
        self.input_widget = input_widget
        try:
            selected = datetime.fromisoformat(input_widget.get()).date()
        except ValueError:
            selected = date.today()
        self.selected = selected
        self.year, self.month = selected.year, selected.month
        self.title("Choose a date")
        self.resizable(False, False)
        self.configure(bg="#f7fafc")
        self.transient(input_widget.winfo_toplevel())
        self._build()
        self.update_idletasks()
        self.geometry(f"+{input_widget.winfo_rootx()}+{input_widget.winfo_rooty() + input_widget.winfo_height() + 4}")
        self.focus_set()

    def _build(self) -> None:
        self.header = tk.Frame(self, bg="#12304a", padx=8, pady=7)
        self.header.pack(fill="x")
        tk.Button(self.header, text="<", command=lambda: self._move(-1), relief="flat", bg="#12304a", fg="white", activebackground="#1d4c73", activeforeground="white", width=3).pack(side="left")
        self.month_label = tk.Label(self.header, font=("Segoe UI", 10, "bold"), bg="#12304a", fg="white")
        self.month_label.pack(side="left", expand=True)
        tk.Button(self.header, text=">", command=lambda: self._move(1), relief="flat", bg="#12304a", fg="white", activebackground="#1d4c73", activeforeground="white", width=3).pack(side="right")
        self.days = tk.Frame(self, bg="#f7fafc", padx=8, pady=8)
        self.days.pack()
        self._render_month()

    def _move(self, delta: int) -> None:
        month = self.month + delta
        if month == 0:
            self.year, self.month = self.year - 1, 12
        elif month == 13:
            self.year, self.month = self.year + 1, 1
        else:
            self.month = month
        self._render_month()

    def _render_month(self) -> None:
        for child in self.days.winfo_children():
            child.destroy()
        self.month_label.configure(text=f"{calendar.month_name[self.month]} {self.year}")
        for col, name in enumerate(("Sun", "Mon", "Tue", "Wed", "Thu", "Fri", "Sat")):
            tk.Label(self.days, text=name, width=4, font=("Segoe UI", 8, "bold"), bg="#f7fafc", fg="#64748b").grid(row=0, column=col, padx=1, pady=(0, 3))
        weeks = calendar.Calendar(firstweekday=6).monthdatescalendar(self.year, self.month)
        for row, week in enumerate(weeks, start=1):
            for col, day in enumerate(week):
                active = day.month == self.month
                selected = day == self.selected
                button = tk.Button(
                    self.days, text=str(day.day), width=4, relief="flat", font=("Segoe UI", 9),
                    bg="#087e6b" if selected else "#f7fafc", fg="white" if selected else ("#17324d" if active else "#b3bdc9"),
                    activebackground="#c9eee6", command=lambda chosen=day: self._choose(chosen),
                )
                button.grid(row=row, column=col, padx=1, pady=1)

    def _choose(self, chosen: date) -> None:
        self.input_widget.set(chosen.isoformat())
        self.destroy()


class DateInput(ttk.Frame):
    """Entry with a calendar button; mirrors the Entry methods used in the forms."""
    def __init__(self, parent: tk.Misc, width: int = 18):
        super().__init__(parent)
        self.entry = ttk.Entry(self, width=width)
        self.entry.pack(side="left")
        ttk.Button(self, text="Pick", style="Accent.TButton", command=lambda: CalendarPopup(self)).pack(side="left", padx=(5, 0))

    def get(self) -> str:
        return self.entry.get()

    def set(self, value: str) -> None:
        self.entry.delete(0, "end")
        self.entry.insert(0, value)

    def delete(self, start, end=None) -> None:
        self.entry.delete(start, end)

    def insert(self, index, value) -> None:
        self.entry.insert(index, value)


class Database:
    def __init__(self, path: Path):
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON")
        self._create_schema()
        self._seed_settings()
        self._migrate_schema()

    def _create_schema(self) -> None:
        self.conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT NOT NULL
            );
            CREATE TABLE IF NOT EXISTS trips (
                id INTEGER PRIMARY KEY,
                event_date TEXT NOT NULL,
                driver TEXT NOT NULL,
                kms REAL NOT NULL CHECK(kms > 0),
                price_cents REAL CHECK(price_cents IS NULL OR price_cents >= 0),
                fuel_efficiency REAL,
                attendees TEXT NOT NULL,
                notes TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS expenses (
                id INTEGER PRIMARY KEY,
                event_date TEXT NOT NULL,
                category TEXT NOT NULL,
                description TEXT NOT NULL,
                amount REAL NOT NULL CHECK(amount > 0),
                payer TEXT NOT NULL,
                attendees TEXT NOT NULL,
                notes TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            CREATE TABLE IF NOT EXISTS payments (
                id INTEGER PRIMARY KEY,
                event_date TEXT NOT NULL,
                payer TEXT NOT NULL,
                payee TEXT NOT NULL,
                amount REAL NOT NULL CHECK(amount > 0),
                notes TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
            );
            """
        )
        self.conn.commit()

    def _migrate_schema(self) -> None:
        """Keep databases from early app versions usable without changing old math."""
        columns = {row["name"]: row for row in self.conn.execute("PRAGMA table_info(trips)")}
        if "fuel_efficiency" not in columns:
            self.conn.execute("ALTER TABLE trips ADD COLUMN fuel_efficiency REAL")
            columns = {row["name"]: row for row in self.conn.execute("PRAGMA table_info(trips)")}
        # The first version required a price. Rebuild this one table so a trip can
        # be recorded first and priced later without inventing a temporary price.
        if columns["price_cents"]["notnull"]:
            self.conn.executescript(
                """
                ALTER TABLE trips RENAME TO trips_legacy;
                CREATE TABLE trips (
                    id INTEGER PRIMARY KEY,
                    event_date TEXT NOT NULL,
                    driver TEXT NOT NULL,
                    kms REAL NOT NULL CHECK(kms > 0),
                    price_cents REAL CHECK(price_cents IS NULL OR price_cents >= 0),
                    fuel_efficiency REAL,
                    attendees TEXT NOT NULL,
                    notes TEXT NOT NULL DEFAULT '',
                    created_at TEXT NOT NULL DEFAULT CURRENT_TIMESTAMP
                );
                INSERT INTO trips(id, event_date, driver, kms, price_cents, fuel_efficiency, attendees, notes, created_at)
                SELECT id, event_date, driver, kms, price_cents, fuel_efficiency, attendees, notes, created_at
                FROM trips_legacy;
                DROP TABLE trips_legacy;
                """
            )
        for driver, efficiency in self.drivers.items():
            self.conn.execute(
                "UPDATE trips SET fuel_efficiency = ? WHERE driver = ? AND fuel_efficiency IS NULL",
                (efficiency, driver),
            )
        # Correct the legacy Food entry once. The user confirmed that Uday paid it.
        if not self.get_setting("food_split_payer_corrected_v2"):
            self.conn.execute(
                "UPDATE expenses SET payer = ?, notes = ? WHERE description = ? AND category = ?",
                ("Uday", "Confirmed: Uday paid the $88.52 food expense; it is split equally with Gurpreet.", "Food shared expense (imported from Cost Split)", "Food"),
            )
            self.set_setting("food_split_payer_corrected_v2", "1")
        self.conn.commit()

    def _seed_settings(self) -> None:
        if self.get_setting("people") is None:
            self.set_setting("people", json.dumps(DEFAULT_PEOPLE))
        if self.get_setting("drivers") is None:
            self.set_setting("drivers", json.dumps(DEFAULT_DRIVERS))

    def get_setting(self, key: str, default: str | None = None) -> str | None:
        row = self.conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
        return row["value"] if row else default

    def set_setting(self, key: str, value: str) -> None:
        self.conn.execute(
            "INSERT INTO settings(key, value) VALUES (?, ?) "
            "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
            (key, value),
        )
        self.conn.commit()

    @property
    def people(self) -> list[str]:
        return json.loads(self.get_setting("people", "[]") or "[]")

    @property
    def drivers(self) -> dict[str, float]:
        return {name: float(rate) for name, rate in json.loads(self.get_setting("drivers", "{}") or "{}").items()}

    def event_count(self) -> int:
        return sum(
            self.conn.execute(f"SELECT COUNT(*) FROM {table}").fetchone()[0]
            for table in ("trips", "expenses", "payments")
        )

    def add_trip(self, event_date: str, driver: str, kms: float, price_cents: float | None, attendees: list[str], notes: str = "") -> None:
        efficiency = self.drivers.get(driver)
        if efficiency is None:
            raise ValueError(f"No fuel efficiency is set for {driver}.")
        self.conn.execute(
            "INSERT INTO trips(event_date, driver, kms, price_cents, fuel_efficiency, attendees, notes) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (event_date, driver, kms, price_cents, efficiency, json.dumps(attendees), notes.strip()),
        )
        self.conn.commit()

    def update_trip(self, event_id: int, event_date: str, driver: str, kms: float, price_cents: float | None, fuel_efficiency: float, attendees: list[str], notes: str = "") -> None:
        self.conn.execute(
            "UPDATE trips SET event_date=?, driver=?, kms=?, price_cents=?, fuel_efficiency=?, attendees=?, notes=? WHERE id=?",
            (event_date, driver, kms, price_cents, fuel_efficiency, json.dumps(attendees), notes.strip(), event_id),
        )
        self.conn.commit()

    def update_pending_trip_prices(self, price_cents: float) -> int:
        pending = self.conn.execute("SELECT COUNT(*) FROM trips WHERE price_cents IS NULL").fetchone()[0]
        if not pending:
            return 0
        self.conn.execute("UPDATE trips SET price_cents = ? WHERE price_cents IS NULL", (price_cents,))
        self.conn.commit()
        return int(pending)

    def add_expense(self, event_date: str, category: str, description: str, amount: float, payer: str, attendees: list[str], notes: str = "") -> None:
        self.conn.execute(
            "INSERT INTO expenses(event_date, category, description, amount, payer, attendees, notes) VALUES (?, ?, ?, ?, ?, ?, ?)",
            (event_date, category.strip(), description.strip(), round_money(amount), payer, json.dumps(attendees), notes.strip()),
        )
        self.conn.commit()

    def update_expense(self, event_id: int, event_date: str, category: str, description: str, amount: float, payer: str, attendees: list[str], notes: str = "") -> None:
        self.conn.execute(
            "UPDATE expenses SET event_date=?, category=?, description=?, amount=?, payer=?, attendees=?, notes=? WHERE id=?",
            (event_date, category.strip(), description.strip(), round_money(amount), payer, json.dumps(attendees), notes.strip(), event_id),
        )
        self.conn.commit()

    def add_payment(self, event_date: str, payer: str, payee: str, amount: float, notes: str = "") -> None:
        self.conn.execute(
            "INSERT INTO payments(event_date, payer, payee, amount, notes) VALUES (?, ?, ?, ?, ?)",
            (event_date, payer, payee, round_money(amount), notes.strip()),
        )
        self.conn.commit()

    def update_payment(self, event_id: int, event_date: str, payer: str, payee: str, amount: float, notes: str = "") -> None:
        self.conn.execute(
            "UPDATE payments SET event_date=?, payer=?, payee=?, amount=?, notes=? WHERE id=?",
            (event_date, payer, payee, round_money(amount), notes.strip(), event_id),
        )
        self.conn.commit()

    def delete_event(self, kind: str, event_id: int) -> None:
        table = {"Trip": "trips", "Expense": "expenses", "Payment": "payments"}.get(kind)
        if not table:
            raise ValueError("Unknown event type.")
        self.conn.execute(f"DELETE FROM {table} WHERE id = ?", (event_id,))
        self.conn.commit()

    def import_original_workbook_once(self, workbook: Path = SOURCE_WORKBOOK) -> tuple[bool, str]:
        """Import the supplied workbook exactly once, with special legacy rows made explicit."""
        if self.get_setting("spreadsheet_imported"):
            return False, "The original spreadsheet has already been imported."
        if not workbook.exists():
            return False, "No 407.xlsx was found beside the app."
        if self.event_count():
            return False, "This database already has entries, so an import was not applied."
        try:
            from openpyxl import load_workbook
        except ImportError:
            return False, "Spreadsheet import needs openpyxl. Install it with: py -m pip install openpyxl"

        try:
            book = load_workbook(workbook, data_only=True)
            sheet = book["Tracking"]
            trips = tolls = 0
            for row in range(2, sheet.max_row + 1):
                event_date, driver, kms, people_count, share, gas_price = [sheet.cell(row, col).value for col in range(1, 7)]
                if not event_date or not driver or not kms or not people_count:
                    continue
                if not isinstance(event_date, datetime):
                    continue
                try:
                    kms = float(kms)
                    people_count = int(people_count)
                    gas_price = float(gas_price)
                except (TypeError, ValueError):
                    continue
                attendees = ["Uday", "Gurpreet"] if people_count == 2 else list(DEFAULT_PEOPLE)
                if kms == 407:
                    total = round(float(share) * people_count, 2)
                    self.add_expense(event_date.date().isoformat(), "Toll", "407 toll (imported from spreadsheet)", total, str(driver), attendees, "Source row marked 407 km; imported as the hard-coded $73.04 toll.")
                    tolls += 1
                else:
                    self.add_trip(event_date.date().isoformat(), str(driver), kms, gas_price, attendees, "Imported from 407.xlsx")
                    trips += 1

            # The Cost Split area has a May 26 food cost of $88.52, split between
            # Uday and Gurpreet. Uday confirmed that he paid it.
            source_date = sheet.cell(9, 10).value
            uday_balance = sheet.cell(9, 12).value
            gurpreet_balance = sheet.cell(9, 13).value
            if isinstance(source_date, datetime) and uday_balance and gurpreet_balance:
                amount = round(abs(float(uday_balance)) + abs(float(gurpreet_balance)), 2)
                self.add_expense(source_date.date().isoformat(), "Food", "Food shared expense (imported from Cost Split)", amount, "Uday", ["Uday", "Gurpreet"], "Confirmed: Uday paid the $88.52 food expense; it is split equally with Gurpreet.")
            self.set_setting("spreadsheet_imported", datetime.now().isoformat(timespec="seconds"))
            return True, f"Imported {trips} fuel trips, {tolls} 407 toll, and the May 26 food split."
        except Exception as exc:
            self.conn.rollback()
            return False, f"Import did not complete: {exc}"

    def all_events(self) -> list[dict]:
        events: list[dict] = []
        for row in self.conn.execute("SELECT * FROM trips"):
            rate = row["fuel_efficiency"] if row["fuel_efficiency"] is not None else self.drivers.get(row["driver"], 0.0)
            price_cents = row["price_cents"]
            total = None if price_cents is None else round_money(row["kms"] * rate * price_cents / 10000)
            attendees = json.loads(row["attendees"])
            shares = {} if total is None else split_cents(total, attendees, row["driver"])
            share_values = list(shares.values())
            events.append({
                "id": row["id"], "kind": "Trip", "display_type": "Fuel trip" if total is not None else "Fuel trip - price pending", "date": row["event_date"], "description": f"{row['kms']:g} km driven by {row['driver']}",
                "payer": row["driver"], "attendees": attendees, "total": total, "shares": shares, "share": share_values[0] if len(set(share_values)) == 1 else None, "notes": row["notes"],
                "price_cents": price_cents, "fuel_efficiency": rate, "detail": f"{row['kms']:g} km x {rate:g} L/100 km x {price_cents:g} cents/L" if total is not None else f"{row['kms']:g} km at {rate:g} L/100 km; gas price not entered yet",
            })
        for row in self.conn.execute("SELECT * FROM expenses"):
            attendees = json.loads(row["attendees"])
            total = round_money(row["amount"])
            shares = split_cents(total, attendees, row["payer"])
            share_values = list(shares.values())
            events.append({
                "id": row["id"], "kind": "Expense", "display_type": f"Shared {row['category'].lower()} expense", "category": row["category"], "date": row["event_date"], "description": f"{row['category']}: {row['description']}",
                "payer": row["payer"], "attendees": attendees, "total": total, "shares": shares, "share": share_values[0] if len(set(share_values)) == 1 else None, "notes": row["notes"], "detail": row["category"],
            })
        for row in self.conn.execute("SELECT * FROM payments"):
            events.append({
                "id": row["id"], "kind": "Payment", "display_type": "Payment sent", "date": row["event_date"], "description": f"{row['payer']} paid {row['payee']}",
                "payer": row["payer"], "payee": row["payee"], "attendees": [], "total": round_money(row["amount"]), "shares": {}, "share": None, "notes": row["notes"], "detail": "Settlement payment",
            })
        return sorted(events, key=lambda event: (event["date"], event["kind"], event["id"]), reverse=True)

    def ledger(self) -> tuple[dict[str, float], list[dict]]:
        # Keep balances as whole cents internally. This avoids a one-cent or
        # floating-point mismatch after many otherwise-correct equal splits.
        balances_cents = defaultdict(int, {person: 0 for person in self.people})
        lines: list[dict] = []
        for event in self.all_events():
            if event["total"] is None:
                continue
            total_cents = int(Decimal(str(event["total"])) * 100)
            if event["kind"] == "Payment":
                balances_cents[event["payer"]] += total_cents
                balances_cents[event["payee"]] -= total_cents
                lines.append({**event, "person": event["payer"], "amount": event["total"], "reason": f"Paid {event['payee']}"})
                lines.append({**event, "person": event["payee"], "amount": -event["total"], "reason": f"Received from {event['payer']}"})
                continue
            for attendee in event["attendees"]:
                balances_cents[attendee] -= int(Decimal(str(event["shares"][attendee])) * 100)
                lines.append({**event, "person": attendee, "amount": -event["shares"][attendee], "reason": "Share of cost"})
            balances_cents[event["payer"]] += total_cents
            lines.append({**event, "person": event["payer"], "amount": event["total"], "reason": "Paid upfront"})
        return {person: cents / 100 for person, cents in balances_cents.items()}, lines

    def settlements(self) -> list[tuple[str, str, float]]:
        balances, _ = self.ledger()
        debtors = [[name, -amount] for name, amount in balances.items() if amount < -0.005]
        creditors = [[name, amount] for name, amount in balances.items() if amount > 0.005]
        suggestions = []
        while debtors and creditors:
            debtor, debt = debtors[0]
            creditor, credit = creditors[0]
            paid = min(debt, credit)
            suggestions.append((debtor, creditor, paid))
            debtors[0][1] -= paid
            creditors[0][1] -= paid
            if debtors[0][1] < 0.005:
                debtors.pop(0)
            if creditors and creditors[0][1] < 0.005:
                creditors.pop(0)
        return suggestions

    def export_csv(self, destination: Path) -> int:
        balances, lines = self.ledger()
        with destination.open("w", newline="", encoding="utf-8-sig") as file:
            writer = csv.writer(file)
            writer.writerow(["Date", "Type", "Description", "Person", "Ledger amount", "Reason", "Paid by", "Total event cost", "Notes"])
            for line in sorted(lines, key=lambda row: (row["date"], row["id"]), reverse=True):
                writer.writerow([line["date"], line["kind"], line["description"], line["person"], round(line["amount"], 2), line["reason"], line["payer"], round(line["total"], 2), line["notes"]])
            writer.writerow([])
            writer.writerow(["Balance check", "", "", "", round(sum(balances.values()), 2)])
        return len(lines)

    def export_xlsx(self, destination: Path) -> int:
        balances, lines = self.ledger()
        workbook = Workbook()
        ledger_sheet = workbook.active
        ledger_sheet.title = "Ledger"
        headers = ["Date", "Type", "Description", "Person", "Ledger amount", "Reason", "Paid by", "Total event cost", "Notes"]
        ledger_sheet.append(headers)
        for cell in ledger_sheet[1]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
        for line in sorted(lines, key=lambda row: (row["date"], row["id"]), reverse=True):
            ledger_sheet.append([line["date"], line["kind"], line["description"], line["person"], round(line["amount"], 2), line["reason"], line["payer"], round(line["total"], 2), line["notes"]])

        balance_sheet = workbook.create_sheet("Balances")
        balance_sheet.append(["Person", "Balance"])
        for cell in balance_sheet[1]:
            font = copy(cell.font)
            font.bold = True
            cell.font = font
        for person, balance in balances.items():
            balance_sheet.append([person, round(balance, 2)])
        balance_sheet.append([])
        balance_sheet.append(["Balance check", round(sum(balances.values()), 2)])

        workbook.save(destination)
        return len(lines)


class PeopleChecks(ttk.Frame):
    def __init__(self, parent: tk.Misc, people: list[str], selected: list[str] | None = None):
        super().__init__(parent)
        self.vars: dict[str, tk.BooleanVar] = {}
        for person in people:
            value = tk.BooleanVar(value=person in (selected or []))
            self.vars[person] = value
            ttk.Checkbutton(self, text=person, variable=value).pack(side="left", padx=(0, 12))

    def selected(self) -> list[str]:
        return [person for person, value in self.vars.items() if value.get()]


class GasTrackerApp(tk.Tk):
    def __init__(self, db: Database):
        super().__init__()
        self.db = db
        self._publishing = False
        self._publish_pending = False
        self.title(APP_NAME)
        self.geometry("1180x760")
        self.minsize(900, 610)
        self.configure(bg="#eef3f8")
        self._setup_style()
        self._build_ui()
        self.refresh_all()

    def _setup_style(self) -> None:
        style = ttk.Style(self)
        style.theme_use("clam")
        style.configure("TFrame", background="#eef3f8")
        style.configure("TLabel", background="#eef3f8", foreground="#243447", font=("Segoe UI", 10))
        style.configure("TNotebook", background="#eef3f8", borderwidth=0)
        style.configure("TNotebook.Tab", padding=(18, 10), background="#dce5f0", foreground="#385069", font=("Segoe UI", 10, "bold"))
        style.map("TNotebook.Tab", background=[("selected", "#12304a")], foreground=[("selected", "white")])
        style.configure("Title.TLabel", font=("Segoe UI", 20, "bold"), foreground="#12304a", background="#eef3f8")
        style.configure("Sub.TLabel", font=("Segoe UI", 10), foreground="#66798e", background="#eef3f8")
        style.configure("Card.TFrame", background="white", relief="solid", borderwidth=1)
        style.configure("CardTitle.TLabel", font=("Segoe UI", 10, "bold"), foreground="#66798e", background="white")
        style.configure("CardValue.TLabel", font=("Segoe UI", 18, "bold"), foreground="#12304a", background="white")
        style.configure("Primary.TButton", font=("Segoe UI", 10, "bold"), padding=(15, 9), background="#087e6b", foreground="white", borderwidth=0)
        style.map("Primary.TButton", background=[("active", "#056454")])
        style.configure("Accent.TButton", font=("Segoe UI", 10, "bold"), padding=(12, 7), background="#12304a", foreground="white", borderwidth=0)
        style.map("Accent.TButton", background=[("active", "#1d4c73")])
        style.configure("TLabelFrame", background="white", bordercolor="#d8e1eb", relief="solid")
        style.configure("TLabelFrame.Label", background="white", foreground="#12304a", font=("Segoe UI", 10, "bold"))
        style.configure("Treeview", font=("Segoe UI", 10), rowheight=34, background="white", fieldbackground="white", bordercolor="#d8e1eb")
        style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"), background="#dce5f0", foreground="#12304a", relief="flat")

    def _build_ui(self) -> None:
        header = tk.Frame(self, bg="#102b43", padx=28, pady=20)
        header.pack(fill="x")
        tk.Label(header, text="GAS / SPLIT", font=("Segoe UI", 21, "bold"), fg="white", bg="#102b43").pack(anchor="w")
        tk.Label(header, text="Track every ride, shared cost, and payment in one clear place.", font=("Segoe UI", 10), fg="#b7cadb", bg="#102b43").pack(anchor="w", pady=(3, 0))
        controls = tk.Frame(header, bg="#102b43")
        controls.pack(side="right", anchor="n")
        self.publish_status = tk.Label(controls, text="Publish idle", font=("Segoe UI", 9), fg="#bdd0df", bg="#102b43")
        self.publish_status.pack(side="right", padx=(0, 8))
        ttk.Button(controls, text="Publish now", style="Accent.TButton", command=self._publish_readonly_snapshot).pack(side="right")
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=24, pady=(16, 22))
        self.dashboard_tab = ttk.Frame(self.notebook, padding=18)
        self.trip_tab = ttk.Frame(self.notebook, padding=22)
        self.expense_tab = ttk.Frame(self.notebook, padding=22)
        self.payment_tab = ttk.Frame(self.notebook, padding=22)
        self.activity_tab = ttk.Frame(self.notebook, padding=18)
        self.settings_tab = ttk.Frame(self.notebook, padding=22)
        self.notebook.add(self.dashboard_tab, text="Dashboard")
        self.notebook.add(self.trip_tab, text="Add Trip")
        self.notebook.add(self.expense_tab, text="Add Expense")
        self.notebook.add(self.payment_tab, text="Record Payment")
        self.notebook.add(self.activity_tab, text="Activity")
        self.notebook.add(self.settings_tab, text="Settings & Backup")
        self._build_dashboard()
        self._build_trip_form()
        self._build_expense_form()
        self._build_payment_form()
        self._build_activity()
        self._build_settings()

    @staticmethod
    def _form_label(parent: tk.Misc, text: str, row: int, column: int = 0) -> None:
        ttk.Label(parent, text=text).grid(row=row, column=column, sticky="w", pady=(7, 2), padx=(0, 12))

    @staticmethod
    def _entry(parent: tk.Misc, width: int = 30) -> ttk.Entry:
        entry = ttk.Entry(parent, width=width)
        return entry

    def _build_dashboard(self) -> None:
        self.dashboard_cards = ttk.Frame(self.dashboard_tab)
        self.dashboard_cards.pack(fill="x")
        self.balance_card_frame = ttk.Frame(self.dashboard_tab)
        self.balance_card_frame.pack(fill="x", pady=(18, 12))
        lower = ttk.Frame(self.dashboard_tab)
        lower.pack(fill="both", expand=True)
        left = ttk.LabelFrame(lower, text="Suggested settlements", padding=14)
        left.pack(side="left", fill="both", expand=True, padx=(0, 9))
        self.settlement_text = tk.Text(left, height=9, wrap="word", font=("Segoe UI", 11), bg="white", relief="flat", state="disabled")
        self.settlement_text.pack(fill="both", expand=True)
        right = ttk.LabelFrame(lower, text="How the balance works", padding=14)
        right.pack(side="left", fill="both", expand=True, padx=(9, 0))
        ttk.Label(right, justify="left", text=(
            "Every cost is split only among the selected people.\n\n"
            "The person who paid is credited for the full cost; each participant is charged their share.\n\n"
            "Positive = this person should receive money.\n"
            "Negative = this person owes money.\n\n"
            "A recorded payment moves money between two people and updates the balance instantly."
        )).pack(anchor="nw")

    def _build_trip_form(self) -> None:
        ttk.Label(self.trip_tab, text="Log a fuel trip", style="Title.TLabel").grid(row=0, column=0, columnspan=3, sticky="w")
        ttk.Label(self.trip_tab, text="Fuel cost is calculated from kilometres, the driver's mileage, and the gas price. Leave the price blank when it is not known yet.", style="Sub.TLabel").grid(row=1, column=0, columnspan=3, sticky="w", pady=(3, 15))
        form = ttk.Frame(self.trip_tab)
        form.grid(row=2, column=0, sticky="nw")
        self.trip_date = DateInput(form); self.trip_date.insert(0, date.today().isoformat())
        self.trip_driver = ttk.Combobox(form, state="readonly", width=27)
        self.trip_kms = self._entry(form); self.trip_kms.insert(0, "110")
        self.trip_price = self._entry(form)
        self.trip_notes = self._entry(form, 42)
        for row, (label, widget) in enumerate((("Date", self.trip_date), ("Driver / payer", self.trip_driver), ("Distance (km)", self.trip_kms), ("Gas price (cents per L, optional)", self.trip_price), ("Note (optional)", self.trip_notes))):
            self._form_label(form, label, row)
            widget.grid(row=row, column=1, sticky="w", pady=(7, 2))
        self._form_label(form, "Who went?", 5)
        self.trip_people_box = PeopleChecks(form, self.db.people, ["Uday", "Gurpreet"])
        self.trip_people_box.grid(row=5, column=1, sticky="w", pady=(7, 2))
        self.trip_preview = ttk.Label(form, text="", foreground="#18765a", font=("Segoe UI", 11, "bold"))
        self.trip_preview.grid(row=6, column=1, sticky="w", pady=(15, 7))
        ttk.Button(form, text="Preview cost", style="Accent.TButton", command=self.preview_trip).grid(row=7, column=1, sticky="w", pady=(3, 5))
        ttk.Button(form, text="Save trip", style="Primary.TButton", command=self.save_trip).grid(row=8, column=1, sticky="w", pady=(8, 0))

    def _build_expense_form(self) -> None:
        ttk.Label(self.expense_tab, text="Log a shared expense", style="Title.TLabel").grid(row=0, column=0, columnspan=3, sticky="w")
        ttk.Label(self.expense_tab, text="Use this for 407 tolls, food, parking, or anything else paid by one person and shared with selected people.", style="Sub.TLabel").grid(row=1, column=0, columnspan=3, sticky="w", pady=(3, 15))
        form = ttk.Frame(self.expense_tab); form.grid(row=2, column=0, sticky="nw")
        self.expense_date = DateInput(form); self.expense_date.insert(0, date.today().isoformat())
        self.expense_category = ttk.Combobox(form, values=["Toll", "Food", "Parking", "Other"], width=27); self.expense_category.set("Toll")
        self.expense_description = self._entry(form, 42)
        self.expense_amount = self._entry(form); self.expense_payer = ttk.Combobox(form, state="readonly", width=27)
        self.expense_notes = self._entry(form, 42)
        fields = (("Date", self.expense_date), ("Category", self.expense_category), ("What was it?", self.expense_description), ("Total amount ($)", self.expense_amount), ("Who paid?", self.expense_payer), ("Note (optional)", self.expense_notes))
        for row, (label, widget) in enumerate(fields):
            self._form_label(form, label, row); widget.grid(row=row, column=1, sticky="w", pady=(7, 2))
        self._form_label(form, "Who shares it?", 6)
        self.expense_people_box = PeopleChecks(form, self.db.people, ["Uday", "Gurpreet"])
        self.expense_people_box.grid(row=6, column=1, sticky="w", pady=(7, 2))
        ttk.Button(form, text="Save expense", style="Primary.TButton", command=self.save_expense).grid(row=7, column=1, sticky="w", pady=(15, 0))

    def _build_payment_form(self) -> None:
        ttk.Label(self.payment_tab, text="Record a payment", style="Title.TLabel").grid(row=0, column=0, columnspan=3, sticky="w")
        ttk.Label(self.payment_tab, text="Enter this only when money actually changed hands. It settles the running balance without changing the original expenses.", style="Sub.TLabel").grid(row=1, column=0, columnspan=3, sticky="w", pady=(3, 15))
        form = ttk.Frame(self.payment_tab); form.grid(row=2, column=0, sticky="nw")
        self.payment_date = DateInput(form); self.payment_date.insert(0, date.today().isoformat())
        self.payment_payer = ttk.Combobox(form, state="readonly", width=27)
        self.payment_payee = ttk.Combobox(form, state="readonly", width=27)
        self.payment_amount = self._entry(form); self.payment_notes = self._entry(form, 42)
        for row, (label, widget) in enumerate((("Date", self.payment_date), ("Paid by", self.payment_payer), ("Paid to", self.payment_payee), ("Amount ($)", self.payment_amount), ("Note / confirmation (optional)", self.payment_notes))):
            self._form_label(form, label, row); widget.grid(row=row, column=1, sticky="w", pady=(7, 2))
        ttk.Button(form, text="Record payment", style="Primary.TButton", command=self.save_payment).grid(row=5, column=1, sticky="w", pady=(15, 0))

    def _build_activity(self) -> None:
        top = ttk.Frame(self.activity_tab); top.pack(fill="x", pady=(0, 10))
        ttk.Label(top, text="Activity", style="Title.TLabel").pack(side="left")
        ttk.Label(top, text="Select an entry to inspect, edit, remove it, or fill in pending gas prices in one pass.", style="Sub.TLabel").pack(side="left", padx=14)
        ttk.Button(top, text="Delete selected", command=self.delete_selected).pack(side="right")
        ttk.Button(top, text="Edit selected", style="Accent.TButton", command=self.edit_selected).pack(side="right", padx=(0, 8))
        ttk.Button(top, text="Fill pending gas prices", style="Primary.TButton", command=self.fill_pending_gas_prices).pack(side="right", padx=(0, 8))
        columns = ("date", "type", "description", "payer", "people", "total", "share")
        self.activity_tree = ttk.Treeview(self.activity_tab, columns=columns, show="headings", selectmode="browse")
        headings = {"date": "Date", "type": "What happened", "description": "Details", "payer": "Paid by / from", "people": "Shared with", "total": "Total", "share": "Each person's share"}
        widths = {"date": 92, "type": 155, "description": 250, "payer": 120, "people": 180, "total": 100, "share": 125}
        for key in columns:
            self.activity_tree.heading(key, text=headings[key]); self.activity_tree.column(key, width=widths[key], anchor="w")
        self.activity_tree.pack(fill="both", expand=True)
        self.activity_tree.bind("<<TreeviewSelect>>", self.show_selected_detail)
        self.activity_tree.bind("<Double-1>", lambda _event: self.edit_selected())
        self.activity_detail = ttk.Label(self.activity_tab, text="Select an entry to see its calculation or note.", wraplength=1000, foreground="#61708a")
        self.activity_detail.pack(fill="x", pady=(10, 0))

    def _build_settings(self) -> None:
        ttk.Label(self.settings_tab, text="Settings & data safety", style="Title.TLabel").grid(row=0, column=0, columnspan=3, sticky="w")
        ttk.Label(self.settings_tab, text="Set your people and driver fuel efficiencies. The data file stays local to this folder.", style="Sub.TLabel").grid(row=1, column=0, columnspan=3, sticky="w", pady=(3, 15))
        form = ttk.Frame(self.settings_tab); form.grid(row=2, column=0, sticky="nw")
        self.people_setting = self._entry(form, 45)
        self.uday_efficiency = self._entry(form); self.gurpreet_efficiency = self._entry(form)
        self._form_label(form, "People (comma-separated)", 0); self.people_setting.grid(row=0, column=1, sticky="w", pady=(7, 2))
        self._form_label(form, "Uday fuel use (L/100 km)", 1); self.uday_efficiency.grid(row=1, column=1, sticky="w", pady=(7, 2))
        self._form_label(form, "Gurpreet fuel use (L/100 km)", 2); self.gurpreet_efficiency.grid(row=2, column=1, sticky="w", pady=(7, 2))
        ttk.Button(form, text="Save settings", style="Primary.TButton", command=self.save_settings).grid(row=3, column=1, sticky="w", pady=(14, 20))
        backups = ttk.LabelFrame(self.settings_tab, text="Backup and export", padding=13)
        backups.grid(row=3, column=0, sticky="nw", pady=(20, 0))
        ttk.Button(backups, text="Create database backup", command=self.backup_database).pack(side="left", padx=(0, 9))
        ttk.Button(backups, text="Export ledger to CSV", command=self.export_csv).pack(side="left")
        ttk.Label(self.settings_tab, text="Tip: keep 407.xlsx as a read-only historical source. This app does not modify it.", foreground="#61708a").grid(row=4, column=0, sticky="w", pady=(15, 0))

    def update_choices(self) -> None:
        drivers = list(self.db.drivers)
        people = self.db.people
        for combo, values in ((self.trip_driver, drivers), (self.expense_payer, people), (self.payment_payer, people), (self.payment_payee, people)):
            current = combo.get(); combo["values"] = values
            if current not in values: combo.set(values[0] if values else "")

    def refresh_all(self) -> None:
        self.update_choices()
        self.refresh_dashboard()
        self.refresh_activity()
        self.people_setting.delete(0, "end"); self.people_setting.insert(0, ", ".join(self.db.people))
        drivers = self.db.drivers
        self.uday_efficiency.delete(0, "end"); self.uday_efficiency.insert(0, str(drivers.get("Uday", "")))
        self.gurpreet_efficiency.delete(0, "end"); self.gurpreet_efficiency.insert(0, str(drivers.get("Gurpreet", "")))

    def _run_publish_task(self, task) -> None:
        threading.Thread(target=task, daemon=True).start()

    def _subprocess_run(self, args, check: bool = False, **kwargs) -> subprocess.CompletedProcess:
        kwargs.setdefault("cwd", DATA_DIR)
        kwargs.setdefault("capture_output", True)
        kwargs.setdefault("text", True)
        kwargs.setdefault("stdin", subprocess.DEVNULL)
        if os.name == "nt":
            startupinfo = subprocess.STARTUPINFO()
            startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
            startupinfo.wShowWindow = subprocess.SW_HIDE
            kwargs["startupinfo"] = startupinfo
            kwargs["creationflags"] = subprocess.CREATE_NO_WINDOW
        result = subprocess.run(args, **kwargs)
        if check and result.returncode != 0:
            raise subprocess.CalledProcessError(result.returncode, result.args, output=result.stdout, stderr=result.stderr)
        return result

    def _git_has_remote(self) -> bool:
        result = self._subprocess_run(["git", "remote"])
        return result.returncode == 0 and bool(result.stdout.strip())

    def _git_has_changes(self) -> bool:
        result = self._subprocess_run(["git", "status", "--porcelain", "--", "docs/index.html", "docs/.nojekyll"])
        return result.returncode == 0 and bool(result.stdout.strip())

    def _git_stage_changes(self) -> bool:
        result = self._subprocess_run(GIT_PUBLISH_COMMAND)
        if result.returncode != 0:
            message = (result.stderr or result.stdout or "Unknown error").strip().splitlines()[0]
            self._set_publish_status(f"Git add failed: {message}", "#f97575")
            return False
        return True

    def _git_commit_changes(self) -> bool:
        result = self._subprocess_run(GIT_COMMIT_COMMAND)
        if result.returncode == 0:
            return True
        output = (result.stderr or result.stdout or "").lower()
        if "nothing to commit" in output:
            return True
        message = (result.stderr or result.stdout or "Unknown error").strip().splitlines()[0]
        self._set_publish_status(f"Git commit failed: {message}", "#f97575")
        return False

    def _git_push_changes(self) -> bool:
        result = self._subprocess_run(GIT_PUSH_COMMAND)
        if result.returncode == 0:
            return True
        message = (result.stderr or result.stdout or "Unknown error").strip().splitlines()[0]
        self._set_publish_status(f"Git push failed: {message}", "#f97575")
        return False

    def _set_publish_status(self, text: str, color: str = "#bdd0df") -> None:
        self.after(0, lambda: self.publish_status.configure(text=text, fg=color))

    def _publish_readonly_snapshot(self) -> None:
        if not (DATA_DIR / ".git").exists() or not self._git_has_remote():
            return
        self._publish_pending = True
        if self._publishing:
            self._set_publish_status("Publish queued…", "#f3b000")
            return
        self._publishing = True
        self._set_publish_status("Publishing snapshot…", "#bdd0df")

        def task() -> None:
            try:
                while True:
                    self._publish_pending = False
                    try:
                        self._create_backup_bundle()
                        self._subprocess_run(PUBLISH_COMMAND, check=True)
                    except subprocess.CalledProcessError:
                        self._set_publish_status("Publish failed", "#f97575")
                        break
                    if self._git_has_changes():
                        if not self._git_stage_changes() or not self._git_commit_changes() or not self._git_push_changes():
                            self._set_publish_status("Publish failed", "#f97575")
                            break
                        self._set_publish_status(f"Published at {datetime.now().strftime('%H:%M:%S')}", "#a8f0c6")
                    else:
                        self._set_publish_status("Snapshot up to date", "#a8f0c6")
                    if not self._publish_pending:
                        break
            finally:
                self._publishing = False

        self._run_publish_task(task)

    def _create_backup_bundle(self) -> Path:
        target_dir = DATA_DIR / "backups"
        target_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        base = target_dir / f"gas_tracker_{timestamp}"
        db_target = base.with_suffix(".db")
        csv_target = base.with_name(base.name + "_ledger.csv")
        xlsx_target = base.with_name(base.name + "_ledger.xlsx")
        self.db.conn.commit()
        shutil.copy2(DB_PATH, db_target)
        self.db.export_csv(csv_target)
        self.db.export_xlsx(xlsx_target)
        return base

    def refresh_dashboard(self) -> None:
        for child in self.dashboard_cards.winfo_children(): child.destroy()
        for child in self.balance_card_frame.winfo_children(): child.destroy()
        events = self.db.all_events(); balances, _ = self.db.ledger()
        pending = sum(event["kind"] == "Trip" and event["total"] is None for event in events)
        summary = [
            ("Entries", str(len(events))),
            ("Total shared costs", money(sum(event["total"] or 0 for event in events if event["kind"] != "Payment"))),
            ("Gas prices pending", str(pending)),
            ("Ledger check", "Balanced" if abs(sum(balances.values())) < 0.01 else "Review needed"),
        ]
        for title, value in summary:
            card = ttk.Frame(self.dashboard_cards, style="Card.TFrame", padding=14); card.pack(side="left", fill="x", expand=True, padx=5)
            ttk.Label(card, text=title, style="CardTitle.TLabel").pack(anchor="w")
            color = "#087e6b" if value == "Balanced" else "#b7472a" if value == "Review needed" else "#12304a"
            ttk.Label(card, text=value, style="CardValue.TLabel", foreground=color).pack(anchor="w", pady=(4, 0))
        for person, balance in balances.items():
            card = ttk.Frame(self.balance_card_frame, style="Card.TFrame", padding=14); card.pack(side="left", fill="x", expand=True, padx=5)
            status = f"should receive {money(balance)}" if balance > 0.005 else f"owes {money(-balance)}" if balance < -0.005 else "settled up"
            color = "#087e6b" if balance > 0.005 else "#b7472a" if balance < -0.005 else "#66798e"
            ttk.Label(card, text=person, style="CardTitle.TLabel").pack(anchor="w")
            ttk.Label(card, text=status, style="CardValue.TLabel", foreground=color).pack(anchor="w", pady=(4, 0))
        text = "Everyone is settled up."
        suggested = self.db.settlements()
        if suggested:
            text = "\n".join(f"- {debtor} pays {creditor} {money(amount)}" for debtor, creditor, amount in suggested)
        if pending:
            text += f"\n\nNote: {pending} fuel trip{'s' if pending != 1 else ''} ha{'ve' if pending != 1 else 's'} no gas price yet and is excluded from the balance."
        self.settlement_text.configure(state="normal"); self.settlement_text.delete("1.0", "end"); self.settlement_text.insert("1.0", text); self.settlement_text.configure(state="disabled")

    def refresh_activity(self) -> None:
        for item in self.activity_tree.get_children(): self.activity_tree.delete(item)
        self.event_by_iid = {}
        for event in self.db.all_events():
            if event["kind"] == "Payment":
                people = f"to {event['payee']}"; share = "-"
            else:
                people = ", ".join(event["attendees"])
                if event["total"] is None:
                    share = "Price pending"
                elif event["share"] is not None:
                    share = money(event["share"])
                else:
                    values = list(event["shares"].values())
                    share = f"{money(min(values))}-{money(max(values))}"
            iid = f"{event['kind']}:{event['id']}"
            self.event_by_iid[iid] = event
            tags = ("pending",) if event["total"] is None else ()
            total = money(event["total"]) if event["total"] is not None else "Price pending"
            self.activity_tree.insert("", "end", iid=iid, values=(event["date"], event["display_type"], event["description"], event["payer"], people, total, share), tags=tags)
        self.activity_tree.tag_configure("pending", background="#fff4d6", foreground="#8a5a00")

    def show_selected_detail(self, _event=None) -> None:
        selected = self.activity_tree.selection()
        if not selected: return
        event = self.event_by_iid[selected[0]]
        extra = f"  Notes: {event['notes']}" if event["notes"] else ""
        if event["kind"] == "Trip":
            if event["total"] is None:
                detail = f"Price pending: this trip has been saved, but it is not included in anyone's balance until the gas price is entered. Open Edit selected to add it later.{extra}"
            else:
                allocations = ", ".join(f"{person} {money(event['shares'][person])}" for person in event["attendees"])
                detail = f"Calculation: {event['detail']} = {money(event['total'])}. Exact split: {allocations}.{extra}"
        elif event["kind"] == "Expense":
            allocations = ", ".join(f"{person} {money(event['shares'][person])}" for person in event["attendees"])
            detail = f"{event['detail']}: {money(event['total'])} split exactly as {allocations}.{extra}"
        else:
            detail = f"Settlement: {event['payer']} paid {event['payee']} {money(event['total'])}.{extra}"
        self.activity_detail.configure(text=detail)

    def preview_trip(self) -> None:
        try:
            driver = self.trip_driver.get(); rate = self.db.drivers[driver]
            attendees = self.trip_people_box.selected()
            if not attendees: raise ValueError("Select at least one person.")
            if driver not in attendees: raise ValueError("The driver must be included as someone who went.")
            price_text = self.trip_price.get().strip()
            if not price_text:
                self.trip_preview.configure(text="Gas price pending - save now, then edit this trip when the price is known.")
                return
            total = round_money(float(self.trip_kms.get()) * rate * float(price_text) / 10000)
            shares = split_cents(total, attendees, driver)
            allocations = ", ".join(f"{person} {money(shares[person])}" for person in attendees)
            self.trip_preview.configure(text=f"Total fuel cost: {money(total)}  |  Exact split: {allocations}")
        except (ValueError, KeyError):
            messagebox.showerror("Check trip", "Enter a valid distance and driver, include the driver in the people who went, and choose at least one person. Gas price may be left blank.")

    def save_trip(self) -> None:
        try:
            attendees = self.trip_people_box.selected()
            driver = self.trip_driver.get()
            if driver not in self.db.drivers or not attendees: raise ValueError("Select a valid driver and at least one person.")
            if driver not in attendees: raise ValueError("The driver must be included as someone who went.")
            price_text = self.trip_price.get().strip()
            price = float(price_text) if price_text else None
            if price is not None and price < 0: raise ValueError("Gas price cannot be negative.")
            self.db.add_trip(parse_date(self.trip_date.get()), driver, float(self.trip_kms.get()), price, attendees, self.trip_notes.get())
        except ValueError as exc:
            messagebox.showerror("Trip not saved", str(exc)); return
        self.trip_notes.delete(0, "end"); self.trip_price.delete(0, "end"); self.trip_preview.configure(text="")
        self.refresh_all(); self._publish_readonly_snapshot(); self.notebook.select(self.dashboard_tab)

    def fill_pending_gas_prices(self) -> None:
        pending_count = self.db.conn.execute("SELECT COUNT(*) FROM trips WHERE price_cents IS NULL").fetchone()[0]
        if not pending_count:
            messagebox.showinfo("No pending prices", "There are no fuel trips waiting for a gas price right now.")
            return

        window = tk.Toplevel(self)
        window.title("Fill pending gas prices")
        window.geometry("560x280")
        window.minsize(520, 250)
        window.configure(bg="#eef3f8")
        window.transient(self)
        window.grab_set()

        header = tk.Frame(window, bg="#102b43", padx=22, pady=15)
        header.pack(fill="x")
        tk.Label(header, text="Fill pending gas prices", font=("Segoe UI", 16, "bold"), fg="white", bg="#102b43").pack(anchor="w")
        tk.Label(header, text=f"Apply one gas price to all {pending_count} pending fuel trip{'s' if pending_count != 1 else ''}.", font=("Segoe UI", 9), fg="#b7cadb", bg="#102b43").pack(anchor="w", pady=(2, 0))

        body = ttk.Frame(window, padding=22)
        body.pack(fill="both", expand=True)
        ttk.Label(body, text="Gas price (cents per L)", font=("Segoe UI", 10, "bold")).grid(row=0, column=0, sticky="w", pady=(0, 6))
        price_entry = self._entry(body, 18)
        price_entry.grid(row=1, column=0, sticky="w")
        ttk.Label(body, text="This updates every saved trip that still has an empty gas price. It does not touch trips that already have a price.", wraplength=500, foreground="#61708a").grid(row=2, column=0, sticky="w", pady=(12, 0))

        actions = ttk.Frame(body)
        actions.grid(row=3, column=0, sticky="w", pady=(22, 0))

        def apply_price() -> None:
            try:
                price_text = price_entry.get().strip()
                if not price_text:
                    raise ValueError("Enter a gas price in cents per litre.")
                price = float(price_text)
                if price < 0:
                    raise ValueError("Gas price cannot be negative.")
                updated = self.db.update_pending_trip_prices(price)
            except ValueError as exc:
                messagebox.showerror("Price not applied", str(exc), parent=window)
                return
            if updated:
                window.destroy()
                self.refresh_all()
                self._publish_readonly_snapshot()
                self.notebook.select(self.activity_tab)
                messagebox.showinfo("Pending prices updated", f"Updated {updated} pending fuel trip{'s' if updated != 1 else ''}.")
            else:
                window.destroy()
                messagebox.showinfo("No pending prices", "There were no pending fuel trips left to update.")

        ttk.Button(actions, text="Apply to all pending trips", style="Primary.TButton", command=apply_price).pack(side="left")
        ttk.Button(actions, text="Cancel", command=window.destroy).pack(side="left", padx=(9, 0))
        price_entry.focus_set()

    def save_expense(self) -> None:
        try:
            attendees = self.expense_people_box.selected(); payer = self.expense_payer.get()
            if not self.expense_description.get().strip(): raise ValueError("Describe the expense.")
            if payer not in self.db.people or not attendees: raise ValueError("Select who paid and at least one person sharing it.")
            self.db.add_expense(parse_date(self.expense_date.get()), self.expense_category.get() or "Other", self.expense_description.get(), float(self.expense_amount.get()), payer, attendees, self.expense_notes.get())
        except ValueError as exc:
            messagebox.showerror("Expense not saved", str(exc)); return
        self.expense_description.delete(0, "end"); self.expense_amount.delete(0, "end"); self.expense_notes.delete(0, "end")
        self.refresh_all(); self._publish_readonly_snapshot(); self.notebook.select(self.dashboard_tab)

    def save_payment(self) -> None:
        try:
            payer, payee = self.payment_payer.get(), self.payment_payee.get()
            if payer not in self.db.people or payee not in self.db.people or payer == payee: raise ValueError("Choose two different people.")
            self.db.add_payment(parse_date(self.payment_date.get()), payer, payee, float(self.payment_amount.get()), self.payment_notes.get())
        except ValueError as exc:
            messagebox.showerror("Payment not saved", str(exc)); return
        self.payment_amount.delete(0, "end"); self.payment_notes.delete(0, "end")
        self.refresh_all(); self._publish_readonly_snapshot(); self.notebook.select(self.dashboard_tab)

    def edit_selected(self) -> None:
        selected = self.activity_tree.selection()
        if not selected:
            messagebox.showinfo("Select an entry", "Choose an activity entry, then select Edit selected.")
            return
        event = self.event_by_iid[selected[0]]
        window = tk.Toplevel(self)
        window.title(f"Edit {event['display_type']}")
        window.geometry("610x560")
        window.minsize(560, 470)
        window.configure(bg="#eef3f8")
        window.transient(self)
        window.grab_set()
        header = tk.Frame(window, bg="#102b43", padx=22, pady=15)
        header.pack(fill="x")
        tk.Label(header, text=f"Edit {event['display_type']}", font=("Segoe UI", 16, "bold"), fg="white", bg="#102b43").pack(anchor="w")
        tk.Label(header, text="Save changes to recalculate the dashboard and balances.", font=("Segoe UI", 9), fg="#b7cadb", bg="#102b43").pack(anchor="w", pady=(2, 0))
        body = ttk.Frame(window, padding=22)
        body.pack(fill="both", expand=True)
        if event["kind"] == "Trip":
            self._edit_trip_window(body, window, event)
        elif event["kind"] == "Expense":
            self._edit_expense_window(body, window, event)
        else:
            self._edit_payment_window(body, window, event)

    def _editor_label(self, parent: tk.Misc, text: str, row: int) -> None:
        ttk.Label(parent, text=text).grid(row=row, column=0, sticky="w", padx=(0, 12), pady=(7, 2))

    def _editor_actions(self, parent: tk.Misc, row: int, save_action, window: tk.Toplevel) -> None:
        actions = ttk.Frame(parent)
        actions.grid(row=row, column=1, sticky="w", pady=(18, 0))
        ttk.Button(actions, text="Save changes", style="Primary.TButton", command=save_action).pack(side="left")
        ttk.Button(actions, text="Cancel", command=window.destroy).pack(side="left", padx=(9, 0))

    def _edit_trip_window(self, form: ttk.Frame, window: tk.Toplevel, event: dict) -> None:
        date_entry = DateInput(form); date_entry.insert(0, event["date"])
        driver = ttk.Combobox(form, state="readonly", values=list(self.db.drivers), width=29); driver.set(event["payer"])
        kms_entry = self._entry(form); kms_entry.insert(0, f"{event['description'].split(' km')[0]}")
        price_entry = self._entry(form)
        if event["price_cents"] is not None: price_entry.insert(0, f"{event['price_cents']:g}")
        efficiency_entry = self._entry(form); efficiency_entry.insert(0, f"{event['fuel_efficiency']:g}")
        note_entry = self._entry(form, 42); note_entry.insert(0, event["notes"])
        fields = (
            ("Date", date_entry), ("Driver / payer", driver), ("Distance (km)", kms_entry),
            ("Gas price (cents per L, optional)", price_entry), ("Fuel use for this trip (L/100 km)", efficiency_entry), ("Note (optional)", note_entry),
        )
        for row, (label, widget) in enumerate(fields):
            self._editor_label(form, label, row); widget.grid(row=row, column=1, sticky="w", pady=(7, 2))
        self._editor_label(form, "Who went?", 6)
        people_box = PeopleChecks(form, self.db.people, event["attendees"])
        people_box.grid(row=6, column=1, sticky="w", pady=(7, 2))

        def save() -> None:
            try:
                attendees = people_box.selected()
                price_text = price_entry.get().strip()
                price = float(price_text) if price_text else None
                if driver.get() not in self.db.drivers or not attendees: raise ValueError("Select a valid driver and at least one person.")
                if driver.get() not in attendees: raise ValueError("The driver must be included as someone who went.")
                if float(kms_entry.get()) <= 0 or float(efficiency_entry.get()) <= 0: raise ValueError("Distance and fuel use must be greater than zero.")
                if price is not None and price < 0: raise ValueError("Gas price cannot be negative.")
                self.db.update_trip(event["id"], parse_date(date_entry.get()), driver.get(), float(kms_entry.get()), price, float(efficiency_entry.get()), attendees, note_entry.get())
            except ValueError as exc:
                messagebox.showerror("Changes not saved", str(exc), parent=window); return
            window.destroy(); self.refresh_all(); self._publish_readonly_snapshot()

        self._editor_actions(form, 7, save, window)

    def _edit_expense_window(self, form: ttk.Frame, window: tk.Toplevel, event: dict) -> None:
        date_entry = DateInput(form); date_entry.insert(0, event["date"])
        category = ttk.Combobox(form, values=["Toll", "Food", "Parking", "Other"], width=29); category.set(event["category"])
        description = self._entry(form, 42); description.insert(0, event["description"].split(": ", 1)[-1])
        amount = self._entry(form); amount.insert(0, f"{event['total']:g}")
        payer = ttk.Combobox(form, state="readonly", values=self.db.people, width=29); payer.set(event["payer"])
        note_entry = self._entry(form, 42); note_entry.insert(0, event["notes"])
        fields = (("Date", date_entry), ("Category", category), ("What was it?", description), ("Total amount ($)", amount), ("Who paid?", payer), ("Note (optional)", note_entry))
        for row, (label, widget) in enumerate(fields):
            self._editor_label(form, label, row); widget.grid(row=row, column=1, sticky="w", pady=(7, 2))
        self._editor_label(form, "Who shares it?", 6)
        people_box = PeopleChecks(form, self.db.people, event["attendees"])
        people_box.grid(row=6, column=1, sticky="w", pady=(7, 2))

        def save() -> None:
            try:
                attendees = people_box.selected()
                if not description.get().strip(): raise ValueError("Describe the expense.")
                if payer.get() not in self.db.people or not attendees: raise ValueError("Select who paid and at least one person sharing it.")
                if float(amount.get()) <= 0: raise ValueError("The amount must be greater than zero.")
                self.db.update_expense(event["id"], parse_date(date_entry.get()), category.get() or "Other", description.get(), float(amount.get()), payer.get(), attendees, note_entry.get())
            except ValueError as exc:
                messagebox.showerror("Changes not saved", str(exc), parent=window); return
            window.destroy(); self.refresh_all(); self._publish_readonly_snapshot()

        self._editor_actions(form, 7, save, window)

    def _edit_payment_window(self, form: ttk.Frame, window: tk.Toplevel, event: dict) -> None:
        date_entry = DateInput(form); date_entry.insert(0, event["date"])
        payer = ttk.Combobox(form, state="readonly", values=self.db.people, width=29); payer.set(event["payer"])
        payee = ttk.Combobox(form, state="readonly", values=self.db.people, width=29); payee.set(event["payee"])
        amount = self._entry(form); amount.insert(0, f"{event['total']:g}")
        note_entry = self._entry(form, 42); note_entry.insert(0, event["notes"])
        fields = (("Date", date_entry), ("Paid by", payer), ("Paid to", payee), ("Amount ($)", amount), ("Note / confirmation (optional)", note_entry))
        for row, (label, widget) in enumerate(fields):
            self._editor_label(form, label, row); widget.grid(row=row, column=1, sticky="w", pady=(7, 2))

        def save() -> None:
            try:
                if payer.get() not in self.db.people or payee.get() not in self.db.people or payer.get() == payee.get(): raise ValueError("Choose two different people.")
                if float(amount.get()) <= 0: raise ValueError("The amount must be greater than zero.")
                self.db.update_payment(event["id"], parse_date(date_entry.get()), payer.get(), payee.get(), float(amount.get()), note_entry.get())
            except ValueError as exc:
                messagebox.showerror("Changes not saved", str(exc), parent=window); return
            window.destroy(); self.refresh_all(); self._publish_readonly_snapshot()

        self._editor_actions(form, 5, save, window)

    def delete_selected(self) -> None:
        selected = self.activity_tree.selection()
        if not selected: messagebox.showinfo("Select an entry", "Choose an activity entry to delete."); return
        event = self.event_by_iid[selected[0]]
        if not messagebox.askyesno("Delete entry", f"Delete this {event['kind'].lower()}? This cannot be undone except by restoring a backup."): return
        self.db.delete_event(event["kind"], event["id"])
        self.activity_detail.configure(text="Select an entry to see its calculation or note.")
        self.refresh_all(); self._publish_readonly_snapshot()

    def save_settings(self) -> None:
        try:
            people = [part.strip() for part in self.people_setting.get().split(",") if part.strip()]
            if len(people) < 2 or len(set(people)) != len(people): raise ValueError("Enter at least two distinct names.")
            # Existing entries retain their names. The spreadsheet uses these two drivers.
            drivers = {"Uday": float(self.uday_efficiency.get()), "Gurpreet": float(self.gurpreet_efficiency.get())}
            if not all(rate > 0 for rate in drivers.values()): raise ValueError("Fuel-use values must be greater than zero.")
            used_names = set()
            for event in self.db.all_events():
                used_names.add(event["payer"]); used_names.update(event["attendees"])
                if event["kind"] == "Payment": used_names.add(event["payee"])
            if not used_names.issubset(people): raise ValueError("Keep the names already used in entries; rename them only before adding data.")
            self.db.set_setting("people", json.dumps(people)); self.db.set_setting("drivers", json.dumps(drivers))
        except ValueError as exc:
            messagebox.showerror("Settings not saved", str(exc)); return
        messagebox.showinfo("Settings saved", "Settings were saved. The new efficiency is used for future trips; historical trip calculations remain unchanged.")
        self.refresh_all()

    def backup_database(self) -> None:
        base = self._create_backup_bundle()
        messagebox.showinfo(
            "Backup created",
            f"Saved:\n{base.with_suffix('.db')}\n{base.with_name(base.name + '_ledger.csv')}\n{base.with_name(base.name + '_ledger.xlsx')}",
        )

    def export_csv(self) -> None:
        name = f"gas_tracker_ledger_{date.today().isoformat()}.csv"
        target = filedialog.asksaveasfilename(title="Export ledger", initialfile=name, defaultextension=".csv", filetypes=[("CSV file", "*.csv")])
        if not target: return
        count = self.db.export_csv(Path(target))
        messagebox.showinfo("Ledger exported", f"Exported {count} ledger lines to:\n{target}")


def main() -> None:
    db = Database(DB_PATH)
    imported, message = db.import_original_workbook_once()
    app = GasTrackerApp(db)
    if imported:
        app.after(350, lambda: messagebox.showinfo("Spreadsheet imported", message + "\n\nThe original 407.xlsx was not modified."))
    app.mainloop()
    db.conn.close()


if __name__ == "__main__":
    main()
